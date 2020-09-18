from flask import Flask, render_template, request, redirect, jsonify
from flask_cors import CORS

#from flask_bootstrap import Bootstrap

import os
import re
import io
import MeCab
import unicodedata
import openpyxl


app = Flask(__name__)
CORS(app)
#app.config['MAX_CONTENT_LENGTH'] = 5 * 1024 * 1024  # 5MB

UPLOAD_FOLDER = '/workspace/'
ALLOWED_EXTENSIONS = {'xlsx'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config["JSON_AS_ASCII"] = False


#@app.route('/')
#def index():
#    return render_template('index.html')


@app.route('/excel', methods=['POST'])
def excel():
    if 'file' not in request.files:
        print('file not found.')
        return redirect('/')

    file = request.files['file']
    filename = file.filename
    print(filename)

    tagger = MeCab.Tagger('-r /etc/mecabrc -E ""')
    results = []

    wb = openpyxl.load_workbook(filename=io.BytesIO(file.read()))
    sheetnames = wb.sheetnames
    sheets = []

    for ws in wb.worksheets:
        sheetname = sheetnames.pop(0)
        cells = []
        for row in ws.rows:
            for cell in row:
                text = unicodedata.normalize('NFKC', str(cell.value))

                if text != "None":
                    result = tagger.parse(text)
                    #print(result)

                    r1 = re.compile(r".*\t名詞,固有名詞,組織,.*")
                    r2 = re.compile(r"\t")
                    r3 = re.compile(r",.*,.*")
                    findresult = r1.findall(result)
                    #print(findresult)
                    for r in findresult:
                        cn = r2.split(r)
                        #print(cn)

                        x = cn[1].replace("名詞,固有名詞,組織,*,*,*,", "")
                        #print(x)
                        y = r3.sub("", x)
                        #print(y)

                        celldata = {
                            "cell": cell.coordinate,
                            "hit": cn[0],
                            "company": y,
                            "text": text
                        }
                        cells.append(celldata)                    
                    
        #print(cells)
        sheetdata = {
            "sheetname": sheetname,
            "cells": cells
        }
        sheets.append(sheetdata)
        #print(sheets)

    filedata = {
        "filename": filename,
        "sheets": sheets
    }
    #print(filedata)
    #print([filedata])
    print(jsonify([filedata]))


    ''' sample json
    a = [
        {
            "filename": "sample.xlsx",
            "sheets": [
                {
                    "sheetname": "testsheet",
                    "cells": [
                        {
                            "cell": "xxxxxx",
                            "hit": "日本語",
                            "company": "日本語株式会社",
                            "text": "日本語が含まれています。"
                        },
                        {
                            "cell": "yyyyy",
                            "hit": "英語",
                            "company": "英語株式会社",
                            "text": "英語が含まれています。"
                        }
                    ]
                }
            ]
        }
    ]
    '''

    return jsonify([filedata])


if __name__ == '__main__':
    app.debug = True
    app.run(host='0.0.0.0', port=5000)
